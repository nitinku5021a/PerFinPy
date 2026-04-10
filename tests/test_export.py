import sys, os, io
from datetime import date
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from app import create_app, db
from app.models import (
    Account,
    JournalEntry,
    TransactionLine,
    MonthlyBudget,
    BudgetEntryAssignment,
    GoalSetting,
    Goal,
    CreditCard,
    TradeSetup,
    TradeJournalEntry
)
from openpyxl import load_workbook

def test_export_transactions():
    app = create_app()

    with app.app_context():
        # fresh DB
        db_path = os.path.join(os.getcwd(), 'accounting.db')
        if os.path.exists(db_path):
            os.remove(db_path)
        db.create_all()

        a1 = Account(code='A1', name='Checking', account_type='Asset')
        a2 = Account(code='E1', name='Groceries', account_type='Expense')
        db.session.add_all([a1, a2])
        db.session.commit()

        je = JournalEntry(entry_date=date(2026, 1, 10), description='Grocery Spend')
        db.session.add(je)
        db.session.flush()
        db.session.add_all([
            TransactionLine(journal_entry_id=je.id, account_id=a2.id, line_type='DEBIT', amount=1200.0, date=date(2026, 1, 10)),
            TransactionLine(journal_entry_id=je.id, account_id=a1.id, line_type='CREDIT', amount=1200.0, date=date(2026, 1, 10))
        ])

        mb = MonthlyBudget(
            month=date(2026, 1, 1),
            budget_amount=20000.0,
            guchi_opening_balance=100.0,
            gunu_opening_balance=200.0
        )
        db.session.add(mb)
        db.session.flush()
        db.session.add(BudgetEntryAssignment(month=date(2026, 1, 1), journal_entry_id=je.id, owner='Guchi'))

        goal_settings = GoalSetting(interest_rate=7.5)
        db.session.add(goal_settings)
        db.session.add(Goal(description='Car fund', target_corpus=500000.0, target_year=2030, current_corpus=150000.0))
        db.session.add(CreditCard(
            card_name='HDFC Millennia',
            holder_name='Guchi',
            card_details='Visa ending 1234',
            features_benefits='Cashback on online spends',
            annual_fee=999.0,
            statement_day=12,
            payment_day=27
        ))
        setup = TradeSetup(name='ORB', start_date=date(2026, 1, 1), is_active=True)
        db.session.add(setup)
        db.session.flush()
        db.session.add(TradeJournalEntry(
            setup_id=setup.id,
            trade_date=date(2026, 1, 10),
            capital_deployed=75000.0,
            pnl_amount=1800.0,
            comment='Good breakout follow-through'
        ))
        db.session.commit()

        client = app.test_client()
        resp = client.get('/transactions/export')
        assert resp.status_code == 200
        data = resp.data
        assert len(data) > 100

        # Load workbook from bytes
        wb = load_workbook(filename=io.BytesIO(data))
        assert 'Transactions' in wb.sheetnames
        ws = wb['Transactions']
        rows = list(ws.rows)
        # header + at least one data row
        assert len(rows) >= 2
        header = [c.value for c in rows[0]]
        assert header == ['Date', 'Debit Account', 'Description', 'Amount', 'Credit Account']

        # Accounts sheet should be present and include Opening Balance
        assert 'Accounts' in wb.sheetnames
        acc_ws = wb['Accounts']
        acc_rows = list(acc_ws.rows)
        assert len(acc_rows) >= 1
        acc_header = [c.value for c in acc_rows[0]]
        assert 'Opening Balance' in acc_header

        # Budget sheets should be present
        assert 'Monthly Budget' in wb.sheetnames
        mb_ws = wb['Monthly Budget']
        mb_rows = list(mb_ws.rows)
        assert len(mb_rows) >= 1
        mb_header = [c.value for c in mb_rows[0]]
        assert mb_header == ['Month', 'Budget Amount', 'Guchi Opening Balance', 'Gunu Opening Balance']

        assert 'Budget Assignments' in wb.sheetnames
        ba_ws = wb['Budget Assignments']
        ba_rows = list(ba_ws.rows)
        assert len(ba_rows) >= 2
        ba_header = [c.value for c in ba_rows[0]]
        assert ba_header == ['Month', 'JE ID', 'Entry Date', 'Description', 'Owner']
        first_assignment = [c.value for c in ba_rows[1]]
        assert first_assignment[0] == '2026-01'
        assert first_assignment[4] == 'Guchi'

        # Goals sheets should be present
        assert 'Goals Settings' in wb.sheetnames
        gs_ws = wb['Goals Settings']
        gs_rows = list(gs_ws.rows)
        assert len(gs_rows) >= 2
        gs_header = [c.value for c in gs_rows[0]]
        assert gs_header == ['Interest Rate %']
        assert gs_rows[1][0].value == 7.5

        assert 'Goals' in wb.sheetnames
        goals_ws = wb['Goals']
        goals_rows = list(goals_ws.rows)
        assert len(goals_rows) >= 2
        goals_header = [c.value for c in goals_rows[0]]
        assert goals_header == ['Goal ID', 'Description', 'Target Corpus', 'Target Year', 'Current Corpus']

        assert 'Credit Cards' in wb.sheetnames
        cc_ws = wb['Credit Cards']
        cc_rows = list(cc_ws.rows)
        assert len(cc_rows) >= 2
        cc_header = [c.value for c in cc_rows[0]]
        assert cc_header == [
            'Card ID',
            'Card Name',
            'Holder Name',
            'Card Details',
            'Features and Benefits',
            'Annual Fee',
            'Statement Date',
            'Payment Date'
        ]
        cc_first_row = [c.value for c in cc_rows[1]]
        assert cc_first_row[1] == 'HDFC Millennia'
        assert cc_first_row[2] == 'Guchi'
        assert cc_first_row[5] == 999
        assert cc_first_row[6] == 12
        assert cc_first_row[7] == 27

        assert 'Trade Setups' in wb.sheetnames
        ts_ws = wb['Trade Setups']
        ts_rows = list(ts_ws.rows)
        assert len(ts_rows) >= 2
        ts_header = [c.value for c in ts_rows[0]]
        assert ts_header == ['Setup ID', 'Setup Name', 'Start Date', 'Is Active']
        ts_first_row = [c.value for c in ts_rows[1]]
        assert ts_first_row[1] == 'ORB'
        assert ts_first_row[2] == '2026-01-01'
        assert ts_first_row[3] is True

        assert 'Trade Journal Entries' in wb.sheetnames
        te_ws = wb['Trade Journal Entries']
        te_rows = list(te_ws.rows)
        assert len(te_rows) >= 2
        te_header = [c.value for c in te_rows[0]]
        assert te_header == ['Entry ID', 'Setup ID', 'Setup Name', 'Trade Date', 'Capital Deployed', 'PnL Amount', 'Comment']
        te_first_row = [c.value for c in te_rows[1]]
        assert te_first_row[2] == 'ORB'
        assert te_first_row[3] == '2026-01-10'
        assert te_first_row[4] == 75000
        assert te_first_row[5] == 1800
        assert te_first_row[6] == 'Good breakout follow-through'

        # spot-check a data row
        first_row = [c.value for c in rows[1]]
        assert first_row[0] is not None and first_row[1] is not None and isinstance(first_row[3], (int, float))
        # debit account path should be prefixed by account type (e.g., 'Asset:...')
        from app.models import AccountType
        debit_path = first_row[1]
        assert isinstance(debit_path, str) and ':' in debit_path and debit_path.split(':')[0] in [t.value for t in AccountType]

