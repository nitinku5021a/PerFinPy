"""
Excel import utility for bulk transaction loading.
Supports hierarchical account format: "TopLevel:MiddleLevel:Account"
"""
from datetime import datetime, date
from app import db
from app.models import (
    Account,
    JournalEntry,
    TransactionLine,
    AccountType,
    MonthlyBudget,
    BudgetEntryAssignment,
    ReminderTask,
    ReminderOccurrence
)
from app.services import snapshots_service


def parse_account_path(path_str, account_type):
    """
    Parse account hierarchy string and resolve/create accounts.
    Format: "TopLevel:MiddleLevel:Account"
    Returns the leaf Account object.
    """
    parts = [p.strip() for p in path_str.split(':')]
    
    if not parts or not parts[0]:
        raise ValueError(f"Invalid account path: {path_str}")
    
    # Ensure we have max 3 levels (1 parent + 1 child + 1 leaf or 1 parent + 1 leaf)
    if len(parts) > 3:
        raise ValueError(f"Account path exceeds max 2 levels: {path_str}")

    # If the first segment is exactly an AccountType name (e.g., "Expense", "Asset"),
    # drop it from the parts so we don't create a top-level account named after the type.
    type_names = [t.value.lower() for t in AccountType]
    if parts and parts[0].strip().lower() in type_names:
        parts = parts[1:]

    if not parts:
        raise ValueError(f"Account path must contain an account name after the type: {path_str}")
    
    parent_account = None
    current_account = None
    
    for i, name in enumerate(parts):
        if not name:
            raise ValueError(f"Empty account name in path: {path_str}")
        
        # Top level - try to find by name (case-insensitive) and account_type
        if i == 0:
            current_account = Account.query.filter(
                db.func.lower(Account.name) == name.lower(),
                Account.parent_id.is_(None),
                Account.account_type == account_type
            ).first()
            if not current_account:
                from uuid import uuid4
                current_account = Account(
                    code=f"__auto__{uuid4().hex[:8]}",
                    name=name,
                    account_type=account_type,
                    parent_id=None
                )
                db.session.add(current_account)
                db.session.flush()
            parent_account = current_account
        else:
            # Child of parent - find by name (case-insensitive) under parent with same account_type
            current_account = Account.query.filter(
                db.func.lower(Account.name) == name.lower(),
                Account.parent_id == parent_account.id,
                Account.account_type == account_type
            ).first()
            if not current_account:
                from uuid import uuid4
                current_account = Account(
                    code=f"__auto__{uuid4().hex[:8]}",
                    name=name,
                    account_type=account_type,
                    parent_id=parent_account.id
                )
                db.session.add(current_account)
                db.session.flush()
            parent_account = current_account
    
    return current_account


def generate_account_code(account_type, level, parent_code=None):
    """
    Generate a unique account code based on type and level.
    """
    type_prefixes = {
        'Asset': '1',
        'Liability': '2',
        'Equity': '3',
        'Income': '4',
        'Expense': '5'
    }
    
    prefix = type_prefixes.get(account_type, '9')
    
    # Count existing accounts of same type/parent to generate unique code
    if parent_code:
        # Child account: parent_code + sequential
        count = Account.query.filter(
            Account.code.startswith(parent_code)
        ).count()
        return f"{parent_code}{count + 1:02d}"
    else:
        # Parent account: type_prefix + sequential
        count = Account.query.filter_by(parent_id=None, account_type=account_type).count()
        return f"{prefix}{count + 1:03d}"


def detect_account_type(path_str, default='Asset'):
    """Try to detect an account type from the top-level name in the path. Falls back to default."""
    top = path_str.split(':')[0].strip().lower()
    if any(k in top for k in ['bank', 'cash', 'asset', 'saving']):
        return 'Asset'
    if any(k in top for k in ['credit', 'card', 'loan', 'liability']):
        return 'Liability'
    if any(k in top for k in ['equity', 'capital', 'owner']):
        return 'Equity'
    if any(k in top for k in ['revenue', 'income', 'sales']):
        return 'Income'
    if any(k in top for k in ['expense', 'expenses', 'cost']):
        return 'Expense'
    return default


def _parse_month_cell(value):
    if value is None or value == '':
        return None
    if isinstance(value, datetime):
        d = value.date()
        return date(d.year, d.month, 1)
    if isinstance(value, date):
        return date(value.year, value.month, 1)
    text = str(value).strip()
    if not text:
        return None
    try:
        parsed = datetime.strptime(text, '%Y-%m')
        return date(parsed.year, parsed.month, 1)
    except Exception:
        parsed = datetime.strptime(text.split()[0], '%Y-%m-%d')
        return date(parsed.year, parsed.month, 1)


def _parse_date_cell(value):
    if value is None or value == '':
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    try:
        return datetime.strptime(text, '%Y-%m-%d').date()
    except Exception:
        from dateutil import parser as _parser
        return _parser.parse(text, dayfirst=True).date()


def _parse_datetime_cell(value):
    if value is None or value == '':
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    text = str(value).strip()
    if not text:
        return None
    try:
        return datetime.strptime(text, '%Y-%m-%d %H:%M:%S')
    except Exception:
        from dateutil import parser as _parser
        return _parser.parse(text, dayfirst=True)


def _parse_bool_cell(value, default=False):
    if value is None or value == '':
        return default
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    text = str(value).strip().lower()
    if text in {'true', 'yes', 'y', '1', 'done'}:
        return True
    if text in {'false', 'no', 'n', '0', 'not done'}:
        return False
    return default


def parse_transaction_row(row, row_num):
    """
    Parse a single transaction row.
    Format: Date, Debit Account, Description, Amount, Credit Account
    Returns tuple: (date, debit_account_obj, description, amount, credit_account_obj)
    """
    try:
        # Row should be: Date, Debit, Desc, Amount, Credit
        if len(row) < 5:
            raise ValueError(f"Row has {len(row)} columns, expected 5 (Date, Debit Account, Description, Amount, Credit Account)")
        
        date_str = str(row[0]).strip()
        debit_path = str(row[1]).strip()
        description = str(row[2]).strip()
        amount_str = str(row[3]).strip()
        credit_path = str(row[4]).strip()
        
        # Parse date (handle both DD-MM-YYYY string and datetime objects from Excel)
        from datetime import date as date_type
        entry_date = None
        
        # If it's already a date/datetime object from Excel
        if isinstance(row[0], (datetime, date_type)):
            entry_date = row[0].date() if hasattr(row[0], 'date') else row[0]
        else:
            # Try to parse as string with explicit day-first handling to avoid MM-DD-YYYY confusion
            date_str = str(row[0]).strip()
            try:
                from dateutil import parser as _parser
                entry_date = _parser.parse(date_str, dayfirst=True).date()
            except Exception:
                # Try YYYY-MM-DD format (ISO) as a fallback
                try:
                    entry_date = datetime.strptime(date_str.split()[0], '%Y-%m-%d').date()
                except Exception:
                    raise ValueError(f"Invalid date format '{date_str}'. Expected DD-MM-YYYY or YYYY-MM-DD")
        
        # Parse amount (allow negative amounts as reversals)
        try:
            amount = float(amount_str)
            if amount == 0:
                raise ValueError(f"Amount must be non-zero, got {amount}")
        except ValueError:
            raise ValueError(f"Invalid amount '{amount_str}'")
        
        # Resolve accounts (create if needed) - attempt to detect types from the path
        debit_type = detect_account_type(debit_path, default='Expense')
        credit_type = detect_account_type(credit_path, default='Asset')

        debit_account = parse_account_path(debit_path, debit_type)
        credit_account = parse_account_path(credit_path, credit_type)
        
        return (entry_date, debit_account, description, amount, credit_account)
    
    except Exception as e:
        raise ValueError(f"Row {row_num}: {str(e)}")


def import_transactions_from_excel(file_stream):
    """
    Import transactions from Excel file.
    Expects columns: Date, Debit Account, Description, Amount, Credit Account
    Returns dict with summary and any errors.
    """
    from openpyxl import load_workbook
    
    results = {
        'success': 0,
        'errors': [],
        'warnings': []
    }
    
    try:
        # Clean up legacy top-level accounts that are named exactly as AccountType (e.g., 'Asset', 'Expense')
        # but have no children and no transaction lines. These often result from older buggy imports.
        type_names = [t.value for t in AccountType]
        removed = []
        for tname in type_names:
            acc = Account.query.filter(
                db.func.lower(Account.name) == tname.lower(),
                Account.parent_id.is_(None)
            ).first()
            if acc:
                # Only remove if it's truly orphaned (no children and no transaction lines pointing to it)
                if not acc.children:
                    tl_count = TransactionLine.query.filter_by(account_id=acc.id).count()
                    if tl_count == 0:
                        db.session.delete(acc)
                        removed.append(tname)
        if removed:
            db.session.commit()
            results['warnings'].append(f"Removed legacy top-level accounts: {', '.join(removed)}")

        workbook = load_workbook(file_stream, data_only=True)

        # Prefer an explicit 'Transactions' sheet; otherwise try to discover a sheet with the expected header
        transactions_ws = None
        expected_header = ['Date', 'Debit Account', 'Description', 'Amount', 'Credit Account']
        if 'Transactions' in workbook.sheetnames:
            transactions_ws = workbook['Transactions']
        else:
            # Try to find a sheet whose first row matches the expected header
            for name in workbook.sheetnames:
                ws_try = workbook[name]
                try:
                    first_row = next(ws_try.iter_rows(min_row=1, max_row=1, values_only=True))
                    first_row_vals = [str(x).strip() if x is not None else '' for x in first_row]
                    if first_row_vals[:5] == expected_header:
                        transactions_ws = ws_try
                        break
                except StopIteration:
                    continue
            # Fallback to active sheet if nothing better found
            if transactions_ws is None:
                transactions_ws = workbook.active

        # If workbook contains an Accounts sheet, read opening balances first
        if 'Accounts' in workbook.sheetnames:
            acc_ws = workbook['Accounts']
            for acc_row_idx, acc_row in enumerate(acc_ws.iter_rows(min_row=2, values_only=True), start=2):
                # Expect columns: Account Path, Opening Balance, Account Type(optional), Code(optional), Description(optional)
                if not acc_row or not acc_row[0]:
                    continue
                acc_path = str(acc_row[0]).strip()
                opening_val = acc_row[1] if len(acc_row) > 1 else None
                try:
                    # Determine explicit type if first segment is an AccountType
                    segs = [s.strip() for s in acc_path.split(':') if s and s.strip()]
                    type_names = [t.value.lower() for t in AccountType]
                    if segs and segs[0].lower() in type_names:
                        account_type = next(t.value for t in AccountType if t.value.lower() == segs[0].lower())
                    else:
                        account_type = detect_account_type(acc_path, default='Asset')

                    acc = parse_account_path(acc_path, account_type)
                    if opening_val is not None and opening_val != '':
                        try:
                            acc.opening_balance = float(opening_val)
                            db.session.add(acc)
                        except Exception:
                            results['warnings'].append(f"Accounts sheet row {acc_row_idx}: invalid opening balance '{opening_val}', skipping")
                except Exception as e:
                    results['warnings'].append(f"Accounts sheet row {acc_row_idx}: {str(e)}")
            # commit account creations/updates before processing transactions
            db.session.commit()

        rows_processed = 0
        for row_idx, row in enumerate(transactions_ws.iter_rows(min_row=2, values_only=True), start=2):
            # Skip empty rows
            if not row or not any(row):
                continue

            try:
                entry_date, debit_account, description, amount, credit_account = parse_transaction_row(row, row_idx)

                # Create journal entry
                je = JournalEntry(
                    entry_date=entry_date,
                    description=description
                )
                db.session.add(je)
                db.session.flush()

                # Normalize negative amount: if amount < 0, treat as reversal (swap debit/credit roles)
                if amount < 0:
                    amount = abs(amount)
                    debit_type = 'CREDIT'
                    credit_type = 'DEBIT'
                else:
                    debit_type = 'DEBIT'
                    credit_type = 'CREDIT'

                # Create debit line
                debit_line = TransactionLine(
                    journal_entry_id=je.id,
                    account_id=debit_account.id,
                    line_type=debit_type,
                    amount=amount,
                    date=entry_date,
                    description=description
                )
                db.session.add(debit_line)

                # Create credit line
                credit_line = TransactionLine(
                    journal_entry_id=je.id,
                    account_id=credit_account.id,
                    line_type=credit_type,
                    amount=amount,
                    date=entry_date,
                    description=description
                )
                db.session.add(credit_line)

                db.session.flush()
                results['success'] += 1
                rows_processed += 1

            except Exception as e:
                results['errors'].append(str(e))
                db.session.rollback()
        
        if results['success'] > 0:
            db.session.commit()
            # After bulk import, rebuild snapshots to guarantee consistency
            snapshots_service.backfill_snapshots(db.session)

        # Import Monthly Budget settings, if present
        if 'Monthly Budget' in workbook.sheetnames:
            mb_ws = workbook['Monthly Budget']
            for row_idx, row in enumerate(mb_ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row or not any(row):
                    continue
                try:
                    month_val = _parse_month_cell(row[0])
                    if not month_val:
                        raise ValueError("Month is required")
                    budget_amount = float(row[1] or 0.0)
                    guchi_opening = float(row[2] or 0.0)
                    gunu_opening = float(row[3] or 0.0)
                    mb = MonthlyBudget.query.filter_by(month=month_val).first()
                    if mb:
                        mb.budget_amount = budget_amount
                        mb.guchi_opening_balance = guchi_opening
                        mb.gunu_opening_balance = gunu_opening
                    else:
                        mb = MonthlyBudget(
                            month=month_val,
                            budget_amount=budget_amount,
                            guchi_opening_balance=guchi_opening,
                            gunu_opening_balance=gunu_opening
                        )
                        db.session.add(mb)
                except Exception as e:
                    results['warnings'].append(f"Monthly Budget row {row_idx}: {str(e)}")

        # Import Budget Assignments, if present
        if 'Budget Assignments' in workbook.sheetnames:
            ba_ws = workbook['Budget Assignments']
            valid_owners = {'Guchi', 'Gunu', 'None'}
            for row_idx, row in enumerate(ba_ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row or not any(row):
                    continue
                try:
                    month_val = _parse_month_cell(row[0])
                    if not month_val:
                        raise ValueError("Month is required")

                    je_id_raw = row[1]
                    entry_date = _parse_date_cell(row[2])
                    description = str(row[3] or '').strip()
                    owner = str(row[4] or 'None').strip() or 'None'
                    if owner not in valid_owners:
                        raise ValueError(f"Invalid owner '{owner}'")

                    je = None
                    if je_id_raw is not None and str(je_id_raw).strip() != '':
                        try:
                            je = JournalEntry.query.get(int(je_id_raw))
                        except Exception:
                            je = None

                    if je is None:
                        q = JournalEntry.query
                        if entry_date:
                            q = q.filter(JournalEntry.entry_date == entry_date)
                        if description:
                            q = q.filter(db.func.lower(JournalEntry.description) == description.lower())
                        je = q.order_by(JournalEntry.id.asc()).first()

                    if je is None:
                        raise ValueError("Journal Entry not found")

                    assignment = BudgetEntryAssignment.query.filter_by(
                        month=month_val,
                        journal_entry_id=je.id
                    ).first()
                    if assignment:
                        assignment.owner = owner
                    else:
                        assignment = BudgetEntryAssignment(
                            month=month_val,
                            journal_entry_id=je.id,
                            owner=owner
                        )
                        db.session.add(assignment)
                except Exception as e:
                    results['warnings'].append(f"Budget Assignments row {row_idx}: {str(e)}")

        # Import Reminder Tasks, if present
        reminder_task_id_map = {}
        if 'Reminders Tasks' in workbook.sheetnames:
            rt_ws = workbook['Reminders Tasks']
            for row_idx, row in enumerate(rt_ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row or not any(row):
                    continue
                try:
                    source_task_id = None
                    if len(row) > 0 and row[0] not in (None, ''):
                        source_task_id = int(row[0])

                    title = str(row[1] or '').strip() if len(row) > 1 else ''
                    if not title:
                        raise ValueError('Title is required')

                    notes = (str(row[2] or '').strip() if len(row) > 2 else '') or None
                    due_day = int(row[3] or 0) if len(row) > 3 else 0
                    if due_day < 1 or due_day > 31:
                        raise ValueError('Due Day Of Month must be between 1 and 31')
                    is_active = _parse_bool_cell(row[4] if len(row) > 4 else None, default=True)

                    task = None
                    if source_task_id is not None:
                        task = ReminderTask.query.get(source_task_id)

                    if task is None:
                        task = (
                            ReminderTask.query
                            .filter(db.func.lower(ReminderTask.title) == title.lower())
                            .filter(ReminderTask.due_day_of_month == due_day)
                            .filter(db.func.coalesce(ReminderTask.notes, '') == (notes or ''))
                            .first()
                        )

                    if task:
                        task.title = title
                        task.notes = notes
                        task.due_day_of_month = due_day
                        task.is_active = is_active
                    else:
                        task = ReminderTask(
                            title=title,
                            notes=notes,
                            due_day_of_month=due_day,
                            is_active=is_active
                        )
                        db.session.add(task)
                        db.session.flush()

                    if source_task_id is not None:
                        reminder_task_id_map[source_task_id] = task.id
                except Exception as e:
                    results['warnings'].append(f"Reminders Tasks row {row_idx}: {str(e)}")

        # Import Reminder Occurrences, if present
        if 'Reminder Occurrences' in workbook.sheetnames:
            ro_ws = workbook['Reminder Occurrences']
            for row_idx, row in enumerate(ro_ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row or not any(row):
                    continue
                try:
                    source_task_id = None
                    if len(row) > 1 and row[1] not in (None, ''):
                        source_task_id = int(row[1])

                    mapped_task_id = None
                    if source_task_id is not None:
                        mapped_task_id = reminder_task_id_map.get(source_task_id)
                        if mapped_task_id is None:
                            existing_task = ReminderTask.query.get(source_task_id)
                            mapped_task_id = existing_task.id if existing_task else None

                    if mapped_task_id is None:
                        raise ValueError('Task not found for occurrence')

                    month_val = _parse_month_cell(row[2] if len(row) > 2 else None)
                    if not month_val:
                        raise ValueError('Month is required')

                    due_date = _parse_date_cell(row[3] if len(row) > 3 else None)
                    if not due_date:
                        raise ValueError('Due Date is required')

                    is_done = _parse_bool_cell(row[4] if len(row) > 4 else None, default=False)
                    done_at = _parse_datetime_cell(row[5] if len(row) > 5 else None)
                    is_removed = _parse_bool_cell(row[6] if len(row) > 6 else None, default=False)
                    removed_at = _parse_datetime_cell(row[7] if len(row) > 7 else None)

                    occurrence = ReminderOccurrence.query.filter_by(
                        reminder_task_id=mapped_task_id,
                        month=month_val
                    ).first()
                    if occurrence:
                        occurrence.due_date = due_date
                        occurrence.is_done = is_done
                        occurrence.done_at = done_at if is_done else None
                        occurrence.is_removed = is_removed
                        occurrence.removed_at = removed_at if is_removed else None
                    else:
                        db.session.add(
                            ReminderOccurrence(
                                reminder_task_id=mapped_task_id,
                                month=month_val,
                                due_date=due_date,
                                is_done=is_done,
                                done_at=done_at if is_done else None,
                                is_removed=is_removed,
                                removed_at=removed_at if is_removed else None
                            )
                        )
                except Exception as e:
                    results['warnings'].append(f"Reminder Occurrences row {row_idx}: {str(e)}")

        db.session.commit()
        
        results['total_rows'] = rows_processed
        
    except Exception as e:
        results['errors'].insert(0, f"Excel parsing error: {str(e)}")
    
    return results
