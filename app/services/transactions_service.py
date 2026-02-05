from datetime import datetime
from app import db
from sqlalchemy import case, select
from app.models import Account, JournalEntry, TransactionLine, AccountType, JournalEntryEditLog, DailyAccountBalance
from app.services.serialization import account_to_dict, entry_to_dict, isoformat_or_none
from app.utils.excel_import import import_transactions_from_excel


def list_transactions(page, period, account_id):
    try:
        from app.services.reports_service import get_period_dates
        start_date, end_date = get_period_dates(period)
    except Exception:
        start_date, end_date = (None, None)

    q = JournalEntry.query
    needs_line_join = False

    if account_id:
        account = Account.query.get(account_id)
        if account:
            descendant_ids = [a.id for a in account.get_all_descendants()] + [account.id]
            q = q.join(TransactionLine).filter(TransactionLine.account_id.in_(descendant_ids))
            needs_line_join = True
        else:
            q = q.join(TransactionLine).filter(TransactionLine.account_id == account_id)
            needs_line_join = True

    if start_date or end_date:
        if not needs_line_join:
            q = q.join(TransactionLine)
            needs_line_join = True
        if start_date:
            q = q.filter(TransactionLine.date >= start_date)
        if end_date:
            q = q.filter(TransactionLine.date <= end_date)

    q = q.order_by(JournalEntry.entry_date.desc())
    if needs_line_join:
        q = q.distinct()

    entries = q.paginate(page=page, per_page=20)

    all_accounts = Account.query.order_by(Account.name).all()
    accounts_for_select = all_accounts

    def _summarize_entry(entry):
        debit_lines = [l for l in entry.transaction_lines if (l.line_type or '').upper() == 'DEBIT']
        credit_lines = [l for l in entry.transaction_lines if (l.line_type or '').upper() == 'CREDIT']

        debit_name = debit_lines[0].account.name if len(debit_lines) == 1 else "Multiple"
        credit_name = credit_lines[0].account.name if len(credit_lines) == 1 else "Multiple"

        amount = 0.0
        if debit_lines:
            amount = sum(l.amount for l in debit_lines)
        elif credit_lines:
            amount = sum(l.amount for l in credit_lines)

        return {
            'id': entry.id,
            'entry_date': isoformat_or_none(entry.entry_date),
            'description': entry.description,
            'reference': entry.reference,
            'debit_account': debit_name,
            'credit_account': credit_name,
            'amount': amount
        }

    def _period_sums():
        type_groups = {
            'Asset': [],
            'Liability': [],
            'Income': [],
            'Expense': []
        }

        accounts = Account.query.filter(Account.account_type.in_(type_groups.keys())).all()
        leaf_accounts = [a for a in accounts if a.is_leaf()]
        account_ids = [a.id for a in leaf_accounts]

        q = db.session.query(
            DailyAccountBalance.account_id,
            db.func.coalesce(db.func.sum(DailyAccountBalance.balance), 0.0)
        ).filter(DailyAccountBalance.account_id.in_(account_ids))
        if start_date:
            q = q.filter(DailyAccountBalance.date >= start_date)
        if end_date:
            q = q.filter(DailyAccountBalance.date <= end_date)
        q = q.group_by(DailyAccountBalance.account_id)
        sums = {row[0]: row[1] for row in q.all()}

        for acc in leaf_accounts:
            val = sums.get(acc.id, 0.0)
            if acc.account_type == 'Income':
                val = abs(val)
            elif acc.account_type == 'Expense':
                val = -abs(val)
            if abs(val) > 0.005:
                type_groups[acc.account_type].append({
                    'account_id': acc.id,
                    'name': acc.name,
                    'value': val
                })

        for key in type_groups:
            type_groups[key] = sorted(type_groups[key], key=lambda x: x['name'].lower())

        return type_groups

    max_entry_date = db.session.query(db.func.max(JournalEntry.entry_date)).scalar()
    min_entry_date = db.session.query(db.func.min(JournalEntry.entry_date)).scalar()

    entry_ids_subq = q.with_entities(JournalEntry.id).subquery()
    entry_ids_select = select(entry_ids_subq.c.id)
    total_amount = (
        db.session.query(db.func.coalesce(db.func.sum(TransactionLine.amount), 0.0))
        .filter(TransactionLine.journal_entry_id.in_(entry_ids_select))
        .filter(db.func.upper(TransactionLine.line_type) == 'DEBIT')
        .scalar()
    )

    account_net_total_all = None
    account_net_total_page = None
    if account_id:
        account = Account.query.get(account_id)
        if account:
            descendant_ids = [a.id for a in account.get_all_descendants()] + [account.id]
        else:
            descendant_ids = [account_id]

        account_net_total_all = (
            db.session.query(db.func.coalesce(db.func.sum(
                case(
                    (db.func.upper(TransactionLine.line_type) == 'DEBIT', TransactionLine.amount),
                    else_=-TransactionLine.amount
                )
            ), 0.0))
            .filter(TransactionLine.journal_entry_id.in_(entry_ids_select))
            .filter(TransactionLine.account_id.in_(descendant_ids))
            .scalar()
        )

        page_entry_ids = [e.id for e in entries.items]
        if page_entry_ids:
            account_net_total_page = (
                db.session.query(db.func.coalesce(db.func.sum(
                    case(
                        (db.func.upper(TransactionLine.line_type) == 'DEBIT', TransactionLine.amount),
                        else_=-TransactionLine.amount
                    )
                ), 0.0))
                .filter(TransactionLine.journal_entry_id.in_(page_entry_ids))
                .filter(TransactionLine.account_id.in_(descendant_ids))
                .scalar()
            )
        else:
            account_net_total_page = 0.0

    return {
        'page': 'transactions_list',
        'entries': [_summarize_entry(e) for e in entries.items],
        'period_sums': _period_sums(),
        'pagination': {
            'page': entries.page,
            'per_page': entries.per_page,
            'pages': entries.pages,
            'total': entries.total,
            'has_next': entries.has_next,
            'has_prev': entries.has_prev
        },
        'total_amount': total_amount,
        'account_net_total_all': account_net_total_all,
        'account_net_total_page': account_net_total_page,
        'period': period,
        'account_id': account_id,
        'accounts_for_select': [account_to_dict(a) for a in accounts_for_select],
        'start_date': isoformat_or_none(start_date),
        'end_date': isoformat_or_none(end_date),
        'max_entry_date': isoformat_or_none(max_entry_date),
        'min_entry_date': isoformat_or_none(min_entry_date)
    }


def get_new_transaction_form_data():
    accounts = Account.query.filter_by(is_active=True).order_by(Account.name).all()
    return {
        'page': 'transactions_new',
        'accounts': [account_to_dict(a) for a in accounts]
    }


def create_transaction(form):
    try:
        entry_date = datetime.strptime(form.get('entry_date') or '', '%Y-%m-%d')
        description = (form.get('description') or '').strip()
        debit_account_id = form.get('debit_account_id')
        credit_account_id = form.get('credit_account_id')
        amount_str = form.get('amount', '0')
        amount = abs(float(amount_str)) if amount_str else 0.0

        if not description:
            raise ValueError('Description is required')
        if not debit_account_id or not credit_account_id:
            raise ValueError('Debit and Credit accounts are required')
        if amount <= 0:
            raise ValueError('Amount must be greater than zero')
        if int(debit_account_id) == int(credit_account_id):
            raise ValueError('Debit and Credit must be different accounts')

        je = JournalEntry(
            entry_date=entry_date,
            description=description,
            reference='',
            notes=''
        )
        db.session.add(je)
        db.session.flush()

        for account_id, line_type in [(debit_account_id, 'DEBIT'), (credit_account_id, 'CREDIT')]:
            tl = TransactionLine(
                journal_entry_id=je.id,
                account_id=int(account_id),
                line_type=line_type,
                amount=amount,
                date=entry_date,
                description=''
            )
            db.session.add(tl)

        db.session.commit()
        return {'entry_id': je.id}
    except Exception:
        db.session.rollback()
        raise


def get_transaction_view(entry_id):
    entry = JournalEntry.query.get_or_404(entry_id)
    return {
        'page': 'transactions_view',
        'entry': entry_to_dict(entry, include_lines=True)
    }


def get_edit_transaction_form_data(entry_id):
    entry = JournalEntry.query.get_or_404(entry_id)
    accounts = Account.query.filter_by(is_active=True).all()
    return {
        'page': 'transactions_edit',
        'entry': entry_to_dict(entry, include_lines=True),
        'accounts': [account_to_dict(a) for a in accounts]
    }


def update_transaction(entry_id, form):
    entry = JournalEntry.query.get_or_404(entry_id)
    try:
        import json
        old_state = {
            'entry': {
                'entry_date': entry.entry_date.isoformat(),
                'description': entry.description,
                'reference': entry.reference,
                'notes': entry.notes
            },
            'lines': [
                {
                    'account_id': l.account_id,
                    'line_type': l.line_type,
                    'amount': l.amount,
                    'description': l.description,
                    'date': l.date.isoformat()
                } for l in entry.transaction_lines
            ]
        }

        entry.entry_date = datetime.strptime(form['entry_date'], '%Y-%m-%d').date()
        entry.description = form['description']
        entry.reference = form.get('reference', '')
        entry.notes = form.get('notes', '')

        for l in list(entry.transaction_lines):
            db.session.delete(l)
        db.session.flush()

        debit_account_id = form.get('debit_account_id')
        credit_account_id = form.get('credit_account_id')

        if debit_account_id and credit_account_id:
            amount_str = form.get('amount', '0')
            amount = abs(float(amount_str)) if amount_str else 0.0

            if int(debit_account_id) == int(credit_account_id):
                raise ValueError('Debit and Credit must be different accounts')
            if amount <= 0:
                raise ValueError('Amount must be greater than zero')

            for account_id, line_type in [(debit_account_id, 'DEBIT'), (credit_account_id, 'CREDIT')]:
                tl = TransactionLine(
                    journal_entry_id=entry.id,
                    account_id=int(account_id),
                    line_type=line_type,
                    amount=amount,
                    date=entry.entry_date,
                    description=entry.description
                )
                db.session.add(tl)
        else:
            line_count_str = form.get('line_count', '0')
            line_count = int(line_count_str) if line_count_str else 0
            for i in range(line_count):
                account_id = form.get(f'line_{i}_account_id')
                line_type = form.get(f'line_{i}_type')
                amount_str = form.get(f'line_{i}_amount', '0')
                amount = float(amount_str) if amount_str else 0.0
                line_desc = form.get(f'line_{i}_description', '')

                if account_id and amount != 0 and line_type:
                    amt = abs(amount)
                    lt = line_type.upper()
                    if amount < 0:
                        lt = 'CREDIT' if lt == 'DEBIT' else 'DEBIT'
                    tl = TransactionLine(
                        journal_entry_id=entry.id,
                        account_id=int(account_id),
                        line_type=lt,
                        amount=amt,
                        date=entry.entry_date,
                        description=line_desc
                    )
                    db.session.add(tl)

        if not entry.is_balanced():
            db.session.rollback()
            raise ValueError('Journal entry must be balanced (debits = credits)')

        new_state = {
            'entry': {
                'entry_date': entry.entry_date.isoformat(),
                'description': entry.description,
                'reference': entry.reference,
                'notes': entry.notes
            },
            'lines': [
                {
                    'account_id': l.account_id,
                    'line_type': l.line_type,
                    'amount': l.amount,
                    'description': l.description,
                    'date': l.date.isoformat()
                } for l in entry.transaction_lines
            ]
        }

        log = JournalEntryEditLog(
            journal_entry_id=entry.id,
            editor=form.get('editor') or None,
            change_summary="Edited via web",
            old_data=json.dumps(old_state),
            new_data=json.dumps(new_state)
        )
        db.session.add(log)

        db.session.commit()
        return {'entry_id': entry.id}
    except Exception:
        db.session.rollback()
        raise


def list_accounts_data():
    accounts = Account.query.order_by(Account.name).all()
    return {
        'page': 'accounts_list',
        'accounts': [account_to_dict(a) for a in accounts]
    }


def get_new_account_form_data():
    account_types = [t.value for t in AccountType]
    parent_accounts = Account.query.filter(Account.parent_id.is_(None)).order_by(Account.name).all()
    return {
        'page': 'accounts_new',
        'account_types': account_types,
        'parent_accounts': [account_to_dict(a) for a in parent_accounts]
    }


def create_account(form):
    try:
        name = form.get('name', '').strip()
        account_type = form.get('account_type', '').strip()
        parent_id = form.get('parent_id')

        if not name:
            raise ValueError('Account name is required')
        if not account_type:
            raise ValueError('Account type is required')

        valid_types = [t.value for t in AccountType]
        if account_type not in valid_types:
            raise ValueError('Invalid account type')

        parent_id_int = int(parent_id) if parent_id else None
        if parent_id_int:
            parent = Account.query.get(parent_id_int)
            if not parent:
                raise ValueError('Selected parent account not found')
            if parent.account_type != account_type:
                raise ValueError('Parent account type must match selected account type')

        q = Account.query.filter(db.func.lower(Account.name) == name.lower(), Account.account_type == account_type)
        if parent_id_int is None:
            q = q.filter(Account.parent_id.is_(None))
        else:
            q = q.filter(Account.parent_id == parent_id_int)
        if q.first():
            raise ValueError('An account with this name already exists under the same parent')

        from uuid import uuid4
        opening_balance_str = form.get('opening_balance', '').strip()
        try:
            opening_balance = float(opening_balance_str) if opening_balance_str else 0.0
        except ValueError:
            raise ValueError('Invalid opening balance')

        account = Account(
            code=f"__auto__{uuid4().hex[:8]}",
            name=name,
            account_type=account_type,
            description=form.get('description', ''),
            parent_id=parent_id_int,
            opening_balance=opening_balance
        )
        db.session.add(account)
        db.session.commit()
        return {'account_id': account.id}
    except Exception:
        db.session.rollback()
        raise


def get_edit_account_form_data(account_id):
    account = Account.query.get_or_404(account_id)
    account_types = [t.value for t in AccountType]
    parent_accounts = Account.query.filter(Account.parent_id.is_(None)).order_by(Account.name).all()
    return {
        'page': 'accounts_edit',
        'account': account_to_dict(account),
        'account_types': account_types,
        'parent_accounts': [account_to_dict(a) for a in parent_accounts]
    }


def edit_account(account_id, form):
    account = Account.query.get_or_404(account_id)
    try:
        name = form.get('name', '').strip()
        account_type = form.get('account_type', '').strip()
        parent_id = form.get('parent_id')
        description = form.get('description', '')

        if not name:
            raise ValueError('Account name is required')
        if not account_type:
            raise ValueError('Account type is required')
        valid_types = [t.value for t in AccountType]
        if account_type not in valid_types:
            raise ValueError('Invalid account type')

        parent_id_int = int(parent_id) if parent_id else None
        if parent_id_int:
            if parent_id_int == account.id:
                raise ValueError('Parent cannot be the account itself')
            descendants = account.get_all_descendants()
            if any(d.id == parent_id_int for d in descendants):
                raise ValueError('Parent cannot be a descendant of the account')
            parent = Account.query.get(parent_id_int)
            if not parent:
                raise ValueError('Selected parent account not found')
            if parent.account_type != account_type:
                raise ValueError('Parent account type must match selected account type')

        if account_type != account.account_type:
            subtree = [account] + account.get_all_descendants()
            subtree_ids = [a.id for a in subtree]
            tl_count = TransactionLine.query.filter(TransactionLine.account_id.in_(subtree_ids)).count()
            if tl_count > 0:
                raise ValueError('Cannot change account type: account or descendants have transactions')
            for a in subtree:
                a.account_type = account_type

        opening_balance_str = form.get('opening_balance', '').strip()
        try:
            opening_balance = float(opening_balance_str) if opening_balance_str else 0.0
        except ValueError:
            raise ValueError('Invalid opening balance')

        q = Account.query.filter(db.func.lower(Account.name) == name.lower(), Account.account_type == account_type, Account.id != account.id)
        if parent_id_int is None:
            q = q.filter(Account.parent_id.is_(None))
        else:
            q = q.filter(Account.parent_id == parent_id_int)
        if q.first():
            raise ValueError('An account with this name already exists under the same parent')

        account.name = name
        account.account_type = account_type
        account.parent_id = parent_id_int
        account.description = description
        account.opening_balance = opening_balance

        db.session.commit()
        return {'account_id': account.id}
    except Exception:
        db.session.rollback()
        raise


def import_transactions(file_storage):
    if file_storage is None:
        raise ValueError('No file selected')
    if file_storage.filename == '':
        raise ValueError('No file selected')
    if not file_storage.filename.endswith(('.xlsx', '.xls')):
        raise ValueError('File must be Excel format (.xlsx or .xls)')
    return import_transactions_from_excel(file_storage.stream)


def export_transactions(period):
    from io import BytesIO
    from openpyxl import Workbook

    try:
        from app.services.reports_service import get_period_dates
        start_date, end_date = get_period_dates(period)
    except Exception:
        start_date, end_date = (None, None)

    q = JournalEntry.query.order_by(JournalEntry.entry_date.asc())
    if start_date:
        q = q.filter(JournalEntry.entry_date >= start_date)
    if end_date:
        q = q.filter(JournalEntry.entry_date <= end_date)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Transactions'
    ws.append(['Date', 'Debit Account', 'Description', 'Amount', 'Credit Account'])

    complex_ws = wb.create_sheet('Complex Entries')
    complex_ws.append(['JE ID', 'Date', 'Description', 'Account Path', 'Type', 'Amount'])

    accounts_ws = wb.create_sheet('Accounts')
    accounts_ws.append(['Account Path', 'Opening Balance', 'Account Type', 'Code', 'Description'])
    accounts = Account.query.order_by(Account.account_type.asc(), Account.parent_id.asc(), Account.name.asc()).all()

    for a in accounts:
        accounts_ws.append([a.get_export_path(), a.opening_balance or 0.0, a.account_type, a.code or '', a.description or ''])

    for je in q.all():
        debit_lines = [l for l in je.transaction_lines if (l.line_type or '').upper() == 'DEBIT']
        credit_lines = [l for l in je.transaction_lines if (l.line_type or '').upper() == 'CREDIT']

        if len(debit_lines) == 1 and len(credit_lines) == 1:
            d = debit_lines[0]
            c = credit_lines[0]
            ws.append([je.entry_date.strftime('%d-%m-%Y'), d.account.get_export_path(), je.description or '', d.amount, c.account.get_export_path()])
        else:
            for l in je.transaction_lines:
                complex_ws.append([je.id, je.entry_date.strftime('%d-%m-%Y'), je.description or '', l.account.get_export_path(), l.line_type, l.amount])

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream
