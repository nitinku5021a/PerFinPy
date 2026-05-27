import csv
import io
import os
from datetime import datetime

from openpyxl import load_workbook

from app import db
from app.models import InvestmentAccount, InvestmentRow, InvestmentCategory, Account, InvestmentInstrumentMapping
from app.services.serialization import account_to_dict


ALLOWED_CATEGORIES = {c.value for c in InvestmentCategory}


def _clean_cell_value(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    return str(value)


def _parse_number(value):
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if not s:
        return 0.0
    s = s.replace(",", "")
    s = s.replace("₹", "").replace("INR", "").strip()
    try:
        return float(s)
    except Exception:
        return 0.0


def _normalize_headers(raw_headers):
    headers = []
    seen = {}
    for idx, raw in enumerate(raw_headers or []):
        header = ("" if raw is None else str(raw)).strip()
        if not header:
            header = f"Column {idx + 1}"
        base = header
        if base in seen:
            seen[base] += 1
            header = f"{base} ({seen[base]})"
        else:
            seen[base] = 1
        headers.append(header)
    return headers


def _parse_csv_bytes(content_bytes):
    text = content_bytes.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        raise ValueError("CSV is empty.")
    headers = _normalize_headers(rows[0])
    data_rows = []
    for raw in rows[1:]:
        if not raw or all((str(c).strip() == "" for c in raw)):
            continue
        padded = list(raw) + [""] * max(0, len(headers) - len(raw))
        data_rows.append([_clean_cell_value(v) for v in padded[: len(headers)]])
    return headers, data_rows


def _parse_xlsx_bytes(content_bytes):
    wb = load_workbook(io.BytesIO(content_bytes), read_only=True, data_only=True)
    ws = wb.worksheets[0]
    it = ws.iter_rows(values_only=True)
    try:
        header_row = next(it)
    except StopIteration:
        raise ValueError("Excel sheet is empty.")
    headers = _normalize_headers(list(header_row or []))
    data_rows = []
    for row in it:
        if row is None:
            continue
        row_list = list(row)
        if not row_list or all((v is None or str(v).strip() == "") for v in row_list):
            continue
        padded = row_list + [""] * max(0, len(headers) - len(row_list))
        data_rows.append([_clean_cell_value(v) for v in padded[: len(headers)]])
    return headers, data_rows


def parse_investment_file(file_storage):
    filename = file_storage.filename or ""
    _, ext = os.path.splitext(filename.lower())
    content_bytes = file_storage.read()
    if not content_bytes:
        raise ValueError("Uploaded file is empty.")
    if ext == ".csv":
        return _parse_csv_bytes(content_bytes)
    if ext in (".xlsx", ".xlsm"):
        return _parse_xlsx_bytes(content_bytes)
    raise ValueError("Unsupported file type. Please upload .xlsx or .csv.")


def list_investments():
    accounts = (
        InvestmentAccount.query.order_by(InvestmentAccount.category.asc(), InvestmentAccount.name.asc()).all()
    )
    payload = []
    for acc in accounts:
        rows = (
            InvestmentRow.query.filter_by(account_id=acc.id)
            .order_by(InvestmentRow.sort_index.asc(), InvestmentRow.id.asc())
            .all()
        )
        payload.append(
            {
                "id": acc.id,
                "category": acc.category,
                "name": acc.name,
                "headers": acc.headers,
                "rows": [
                    {
                        "id": r.id,
                        "sort_index": r.sort_index,
                        "data": r.data,
                        "mapping_account_id": r.mapping_account_id,
                    }
                    for r in rows
                ],
                "updated_at": acc.updated_at.isoformat() if acc.updated_at else None,
            }
        )

    mappings = (
        InvestmentInstrumentMapping.query.order_by(
            InvestmentInstrumentMapping.category.asc(),
            InvestmentInstrumentMapping.instrument.asc(),
        ).all()
    )
    mapping_payload = [
        {
            "category": m.category,
            "instrument": m.instrument,
            "mapping_account_id": m.mapping_account_id,
        }
        for m in mappings
    ]

    return {"accounts": payload, "combined_mappings": mapping_payload}


def upsert_account_from_table(category, account_name, headers, data_rows):
    if category not in ALLOWED_CATEGORIES:
        raise ValueError(f"Invalid category: {category}")
    account_name = (account_name or "").strip()
    if not account_name:
        raise ValueError("Account name is required.")
    headers = headers or []
    if not headers:
        raise ValueError("No headers found in file.")

    acc = InvestmentAccount.query.filter_by(category=category, name=account_name).first()
    if acc is None:
        acc = InvestmentAccount(category=category, name=account_name)
        db.session.add(acc)

    acc.headers = headers
    acc.updated_at = datetime.utcnow()

    existing_mappings = {}
    if acc.id is not None:
        existing_rows = InvestmentRow.query.filter_by(account_id=acc.id).all()
        for r in existing_rows:
            inst = ((r.data or {}).get("Instrument") or "").strip()
            if inst and r.mapping_account_id:
                existing_mappings[inst] = int(r.mapping_account_id)

    InvestmentRow.query.filter_by(account_id=acc.id).delete()
    db.session.flush()

    for idx, row_cells in enumerate(data_rows or []):
        row_dict = {headers[i]: (row_cells[i] if i < len(row_cells) else "") for i in range(len(headers))}
        r = InvestmentRow(account=acc, sort_index=idx)
        r.data = row_dict
        inst = (row_dict.get("Instrument") or "").strip()
        if inst and inst in existing_mappings:
            r.mapping_account_id = existing_mappings[inst]
        db.session.add(r)

    db.session.commit()
    return {"success": True, "account_id": acc.id}


def create_account(category, name, headers):
    if category not in ALLOWED_CATEGORIES:
        raise ValueError(f"Invalid category: {category}")
    name = (name or "").strip()
    if not name:
        raise ValueError("Account name is required.")
    headers = _normalize_headers([h for h in (headers or []) if str(h).strip() != ""])
    if not headers:
        raise ValueError("At least one header is required.")
    existing = InvestmentAccount.query.filter_by(category=category, name=name).first()
    if existing:
        raise ValueError("Account already exists for this category.")
    acc = InvestmentAccount(category=category, name=name)
    acc.headers = headers
    db.session.add(acc)
    db.session.commit()
    return {"success": True, "account_id": acc.id}


def add_row(account_id, data):
    acc = InvestmentAccount.query.get(int(account_id))
    if not acc:
        raise ValueError("Account not found.")
    headers = acc.headers
    if not headers:
        raise ValueError("Account has no headers.")
    incoming = data or {}
    row_dict = {h: _clean_cell_value(incoming.get(h, "")) for h in headers}
    max_index = (
        db.session.query(db.func.max(InvestmentRow.sort_index)).filter(InvestmentRow.account_id == acc.id).scalar()
    )
    next_index = int(max_index or 0) + 1 if max_index is not None else 0
    r = InvestmentRow(account=acc, sort_index=next_index)
    r.data = row_dict
    db.session.add(r)
    db.session.commit()
    return {"success": True, "row_id": r.id}


def delete_row(row_id):
    row = InvestmentRow.query.get(int(row_id))
    if not row:
        raise ValueError("Row not found.")
    db.session.delete(row)
    db.session.commit()
    return {"success": True}


def delete_account(account_id):
    acc = InvestmentAccount.query.get(int(account_id))
    if not acc:
        raise ValueError("Account not found.")
    db.session.delete(acc)
    db.session.commit()
    return {"success": True}


def set_row_mapping(row_id, mapping_account_id):
    row = InvestmentRow.query.get(int(row_id))
    if not row:
        raise ValueError("Row not found.")

    if mapping_account_id in ("", None):
        row.mapping_account_id = None
        db.session.commit()
        return {"success": True}

    acc = Account.query.get(int(mapping_account_id))
    if not acc or not acc.is_leaf() or acc.account_type != "Asset":
        raise ValueError("Mapping must be an Asset leaf account.")

    row.mapping_account_id = acc.id
    db.session.commit()
    return {"success": True}


def _norm_header(h):
    return (
        str(h or "")
        .lower()
        .replace(".", "")
        .replace(" ", "")
        .replace("_", "")
    )


def _get_instrument_metrics_for_category(category, instrument):
    instrument = (instrument or "").strip()
    if not instrument:
        raise ValueError("Instrument is required.")

    accounts = InvestmentAccount.query.filter_by(category=category).all()
    qty_sum = 0.0
    avg_cost_sum = 0.0
    avg_cost_count = 0
    ltp_sum = 0.0
    ltp_count = 0

    for acc in accounts:
        headers = acc.headers or []
        key_by_norm = {_norm_header(h): h for h in headers}
        inst_key = key_by_norm.get("instrument", "Instrument")
        qty_key = key_by_norm.get("qty", key_by_norm.get("qty.", "Qty."))
        avg_key = key_by_norm.get("avgcost", "Avg. cost")
        ltp_key = key_by_norm.get("ltp", "LTP")

        rows = InvestmentRow.query.filter_by(account_id=acc.id).all()
        for r in rows:
            data = r.data or {}
            inst = (data.get(inst_key) or data.get("Instrument") or "").strip()
            if inst != instrument:
                continue

            qty = _parse_number(data.get(qty_key) or data.get("Qty.") or data.get("Qty") or 0.0)
            qty_sum += qty

            avg_cost = _parse_number(data.get(avg_key) or data.get("Avg. cost"))
            if avg_cost:
                avg_cost_sum += avg_cost
                avg_cost_count += 1

            ltp = _parse_number(data.get(ltp_key) or data.get("LTP"))
            if ltp:
                ltp_sum += ltp
                ltp_count += 1

    avg_cost = (avg_cost_sum / avg_cost_count) if avg_cost_count else 0.0
    ltp = (ltp_sum / ltp_count) if ltp_count else 0.0

    invested = qty_sum * avg_cost
    cur_val = qty_sum * ltp
    pnl = cur_val - invested
    net_chg = (pnl / invested * 100.0) if invested else None

    return {
        "qty": qty_sum,
        "avg_cost": avg_cost,
        "ltp": ltp,
        "invested": invested,
        "cur_val": cur_val,
        "pnl": pnl,
        "net_chg": net_chg,
    }


def set_combined_mapping(category, instrument, mapping_account_id):
    if category not in ALLOWED_CATEGORIES:
        raise ValueError(f"Invalid category: {category}")
    instrument = (instrument or "").strip()
    if not instrument:
        raise ValueError("Instrument is required.")

    m = InvestmentInstrumentMapping.query.filter_by(category=category, instrument=instrument).first()
    if m is None:
        m = InvestmentInstrumentMapping(category=category, instrument=instrument)
        db.session.add(m)

    if mapping_account_id in ("", None):
        m.mapping_account_id = None
        db.session.commit()
        return {"success": True}

    acc = Account.query.get(int(mapping_account_id))
    if not acc or not acc.is_leaf() or acc.account_type != "Asset":
        raise ValueError("Mapping must be an Asset leaf account.")

    m.mapping_account_id = acc.id
    db.session.commit()
    return {"success": True}


def sync_combined_preview(category, instrument, as_of_date):
    if category not in ALLOWED_CATEGORIES:
        raise ValueError(f"Invalid category: {category}")
    instrument = (instrument or "").strip()
    if not instrument:
        raise ValueError("Instrument is required.")

    metrics = _get_instrument_metrics_for_category(category, instrument)
    inv_value = float(metrics["cur_val"] or 0.0)

    m = InvestmentInstrumentMapping.query.filter_by(category=category, instrument=instrument).first()
    mapping_missing = (m is None) or (m.mapping_account_id is None)

    asset = None
    asset_value = 0.0
    if not mapping_missing:
        asset = Account.query.get(int(m.mapping_account_id))
        if not asset or not asset.is_leaf() or asset.account_type != "Asset":
            raise ValueError("Mapped Asset account is invalid.")
        asset_value = float(asset.get_balance(end_date=as_of_date) or 0.0)

    # If no mapping, treat as a new buy: cost = invested, and sync from cost -> current value.
    purchase_amount = float(metrics.get("invested") or 0.0) if mapping_missing else None
    effective_asset_value = float(purchase_amount) if mapping_missing else float(asset_value)
    diff = float(inv_value - effective_asset_value)

    offset_candidates = _find_capital_gain_accounts()
    default_offset = offset_candidates[0] if offset_candidates else None

    direction = "no_change"
    amount = 0.0
    if abs(diff) >= 0.005:
        direction = "gain" if diff > 0 else "loss"
        amount = abs(diff)

    return {
        "category": category,
        "instrument": instrument,
        "combined": {
            **metrics,
        },
        "investment_value": inv_value,
        "mapping_missing": mapping_missing,
        "purchase": None if not mapping_missing else {
            "amount": purchase_amount,
        },
        "asset": None if asset is None else {
            "id": asset.id,
            "name": asset.name,
            "path": asset.get_path(),
            "value": asset_value,
        },
        "effective_asset_value": effective_asset_value,
        "diff": diff,
        "direction": direction,
        "amount": amount,
        "offset_options": [account_to_dict(a) for a in offset_candidates],
        "default_offset_account_id": default_offset.id if default_offset else None,
    }


def sync_combined_post(
    category,
    instrument,
    as_of_date,
    offset_account_id,
    create_parent_id=None,
    purchase_offset_account_id=None,
):
    preview = sync_combined_preview(category, instrument, as_of_date)

    created_entry_ids = []

    if preview.get("mapping_missing"):
        if not create_parent_id:
            raise ValueError("Mapping is empty; please select a parent to auto-create the Asset account.")
        from app.services import transactions_service

        if not purchase_offset_account_id:
            raise ValueError("purchase_offset_account_id is required for new purchases.")
        purchase_offset = Account.query.get(int(purchase_offset_account_id))
        if not purchase_offset or not purchase_offset.is_leaf() or purchase_offset.account_type != "Asset":
            raise ValueError("Purchase offset must be an Asset leaf account (cash/bank/broker).")

        created = transactions_service.create_account(
            {
                "name": instrument,
                "account_type": "Asset",
                "parent_id": str(create_parent_id),
                "opening_balance": "0",
                "description": "Auto-created from Investments Combined mapping",
            }
        )
        new_account_id = int(created.get("account_id"))
        set_combined_mapping(category, instrument, new_account_id)

        purchase_amount = float((preview.get("purchase") or {}).get("amount") or 0.0)
        if purchase_amount <= 0:
            raise ValueError("Purchase amount (cost) is 0; cannot create purchase entry.")

        # Entry 1: Buy at cost (Dr new Asset / Cr cash/bank)
        res_buy = transactions_service.create_transaction(
            {
                "entry_date": as_of_date.isoformat(),
                "description": f"Buy {instrument} @ cost",
                "debit_account_id": str(new_account_id),
                "credit_account_id": str(purchase_offset.id),
                "amount": str(purchase_amount),
            }
        )
        if res_buy.get("entry_id"):
            created_entry_ids.append(res_buy["entry_id"])

        # Rebuild preview after purchase so sync uses cost as the baseline
        preview = sync_combined_preview(category, instrument, as_of_date)

    # If after purchase (or normal flow) there is nothing to sync, stop here.
    if preview["direction"] == "no_change" or preview["amount"] <= 0:
        return {
            "success": True,
            "posted": bool(created_entry_ids),
            "entry_ids": created_entry_ids,
            "message": "No difference to sync.",
        }

    asset_id = int(preview["asset"]["id"])
    offset = Account.query.get(int(offset_account_id))
    if not offset or not offset.is_leaf():
        raise ValueError("Offset account is invalid.")

    amount = float(preview["amount"])
    inv_value = float(preview["investment_value"] or 0.0)
    asset_value = float(preview["asset"]["value"] or 0.0)

    if preview["direction"] == "gain":
        debit_account_id = asset_id
        credit_account_id = int(offset.id)
    else:
        debit_account_id = int(offset.id)
        credit_account_id = asset_id

    description = f"Sync {instrument}: {asset_value:.2f} -> {inv_value:.2f}"

    from app.services import transactions_service

    res_sync = transactions_service.create_transaction(
        {
            "entry_date": as_of_date.isoformat(),
            "description": description,
            "debit_account_id": str(debit_account_id),
            "credit_account_id": str(credit_account_id),
            "amount": str(amount),
        }
    )
    if res_sync.get("entry_id"):
        created_entry_ids.append(res_sync["entry_id"])
    return {"success": True, "posted": True, "entry_ids": created_entry_ids}


def _find_capital_gain_accounts():
    candidates = (
        Account.query.filter_by(is_active=True)
        .filter(Account.account_type.in_(["Income", "Expense"]))
        .order_by(Account.name.asc())
        .all()
    )
    by_name = {a.name: a for a in candidates}
    preferred = []
    for name in ["Equity Capital Gain", "Debt Capital Gain"]:
        if name in by_name and by_name[name].is_leaf():
            preferred.append(by_name[name])
    if preferred:
        return preferred
    return [a for a in candidates if a.is_leaf()]


def sync_preview(row_id, as_of_date):
    row = InvestmentRow.query.get(int(row_id))
    if not row:
        raise ValueError("Row not found.")

    data = row.data or {}
    instrument = (data.get("Instrument") or "").strip() or f"Row {row.id}"
    inv_value = _parse_number(data.get("Cur. val"))

    if not row.mapping_account_id:
        raise ValueError("Please set Mapping for this instrument first.")

    asset = Account.query.get(int(row.mapping_account_id))
    if not asset or not asset.is_leaf() or asset.account_type != "Asset":
        raise ValueError("Mapped Asset account is invalid.")

    asset_value = float(asset.get_balance(end_date=as_of_date) or 0.0)
    diff = float(inv_value - asset_value)

    offset_candidates = _find_capital_gain_accounts()
    default_offset = offset_candidates[0] if offset_candidates else None

    direction = "no_change"
    amount = 0.0
    if abs(diff) >= 0.005:
        direction = "gain" if diff > 0 else "loss"
        amount = abs(diff)

    return {
        "row_id": row.id,
        "instrument": instrument,
        "investment_value": inv_value,
        "asset": {
            "id": asset.id,
            "name": asset.name,
            "path": asset.get_path(),
            "value": asset_value,
        },
        "diff": diff,
        "direction": direction,
        "amount": amount,
        "offset_options": [account_to_dict(a) for a in offset_candidates],
        "default_offset_account_id": default_offset.id if default_offset else None,
    }


def sync_post(row_id, as_of_date, offset_account_id):
    preview = sync_preview(row_id, as_of_date)
    if preview["direction"] == "no_change" or preview["amount"] <= 0:
        return {"success": True, "posted": False, "message": "No difference to sync."}

    asset_id = int(preview["asset"]["id"])
    offset = Account.query.get(int(offset_account_id))
    if not offset or not offset.is_leaf():
        raise ValueError("Offset account is invalid.")

    amount = float(preview["amount"])
    instrument = preview["instrument"]
    inv_value = preview["investment_value"]
    asset_value = preview["asset"]["value"]

    if preview["direction"] == "gain":
        debit_account_id = asset_id
        credit_account_id = int(offset.id)
        description = f"Sync {instrument}: {asset_value:.2f} -> {inv_value:.2f}"
    else:
        debit_account_id = int(offset.id)
        credit_account_id = asset_id
        description = f"Sync {instrument}: {asset_value:.2f} -> {inv_value:.2f}"

    from app.services import transactions_service

    res = transactions_service.create_transaction(
        {
            "entry_date": as_of_date.isoformat(),
            "description": description,
            "debit_account_id": str(debit_account_id),
            "credit_account_id": str(credit_account_id),
            "amount": str(amount),
        }
    )
    return {"success": True, "posted": True, "entry_id": res.get("entry_id")}
